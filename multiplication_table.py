#! /usr/bin/env python
# multiplication_table.py -- Creates a multiplication table in excel

import sys
import openpyxl
from openpyxl.styles import Font

# A program that takes a number N from the command line
# and creates an NxN multiplication table in an Excel spreadsheet.
try:
    usr_input = int(sys.argv[1])
except ValueError:
    print "Usage: multiplication_table.py [number]"
    print "[number] means an integer only."
    sys.exit()

except IndexError:
    print "Usage: multiplication_table.py [number]"
    print "[number] means an integer only."
    sys.exit()


# Create a workbook
work_book = openpyxl.Workbook()

# Create worksheet
sheet = work_book.active

# Create header row
sheet.cell(row=1, column=1).value = ''
for header_column in range(1, usr_input + 1):
    sheet.cell(row=1, column=header_column + 1).value = header_column

# Create side column
for side_column in range(1, usr_input + 1):
    sheet.cell(row=side_column + 1, column=1).value = side_column

# Make header row and side column bold
bold12font = Font(size=12, bold=True)
for row in range(1, usr_input + 1):
    sheet.cell(row=row + 1, column=1).font = bold12font  # cell A1 is blank
for column in range(1, usr_input + 1):
    sheet.cell(row=1, column=column + 1).font = bold12font

# Create multiplication-table rows and columns.
for row_num in range(1, usr_input + 1):
    for column_ in range(1, usr_input + 1):
        sheet.cell(
            row=row_num + 1, column=column_ + 1).value = column_ * row_num

work_book.save('mult_table.xlsx')